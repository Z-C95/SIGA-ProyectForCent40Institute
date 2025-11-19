# asistencias/views/admin_views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
import csv

from ..models import (
    User, Alumno, Docente, Carrera, Materia, Periodo,
    DocenteMateria, AlumnoMateria, Asistencia
)
from ..permissions import is_admin
from django.db.models import Count, Case, When, IntegerField


# =========================
# Dashboard
# =========================
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    total_docentes = Docente.objects.count()
    total_alumnos = Alumno.objects.count()
    total_materias = Materia.objects.count()
    total_asistencias = Asistencia.objects.count()
    return render(request, "admin/dashboard.html", {
        "total_docentes": total_docentes,
        "total_alumnos": total_alumnos,
        "total_materias": total_materias,
        "total_asistencias": total_asistencias,
    })


# =========================
# Usuarios (lista/crear/editar/eliminar)
# =========================
@login_required
@user_passes_test(is_admin)
def usuarios_lista(request):
    # Parámetros de filtro desde la URL
    q = request.GET.get("q", "").strip()           # texto libre
    rol = request.GET.get("rol", "").strip()       # '', 'ADMIN', 'DOCENTE', 'ALUMNO'
    activo = request.GET.get("activo", "").strip() # '', '1', '0'
    staff = request.GET.get("staff", "").strip()   # '', '1', '0'

    qs = User.objects.all().order_by("username")

    # Búsqueda
    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(email__icontains=q) |
            Q(rol__icontains=q)
        )

    # Filtros
    if rol:
        qs = qs.filter(rol=rol)
    if activo in ("0", "1"):
        qs = qs.filter(is_active=(activo == "1"))
    if staff in ("0", "1"):
        qs = qs.filter(is_staff=(staff == "1"))

    # Paginación
    paginator = Paginator(qs, 10)  # 10 por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    params = request.GET.copy()
    params.pop("page", None)
    querystring = params.urlencode()

    try:
        roles_choices = getattr(User, "Rol").choices  # [('ADMIN','ADMIN'), ...]
        roles = [{"value": "", "label": "Todos"}] + [{"value": v, "label": v} for v, _ in roles_choices]
    except Exception:
        roles = [
            {"value": "", "label": "Todos"},
            {"value": "ADMIN", "label": "ADMIN"},
            {"value": "DOCENTE", "label": "DOCENTE"},
            {"value": "ALUMNO", "label": "ALUMNO"},
        ]

    context = {
        "usuarios": page_obj,
        "query": q,
        "rol": rol,
        "activo": activo,
        "staff": staff,
        "roles": roles,
        "qs": querystring,  # conservar filtros en links de paginación
    }
    return render(request, "usuarios/lista.html", context)


@login_required
@user_passes_test(is_admin)
def usuarios_crear(request):
    from ..forms import CustomUserCreationForm
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado correctamente.")
            return redirect("asistencias:usuarios_lista")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = CustomUserCreationForm()
    return render(request, "usuarios/crear.html", {"form": form})


@login_required
@user_passes_test(is_admin)
def usuarios_editar(request, user_id):
    from ..forms import CustomUserEditForm
    usuario = get_object_or_404(User, id=user_id)
    if request.method == "POST":
        form = CustomUserEditForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario actualizado correctamente.")
            return redirect("asistencias:usuarios_lista")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = CustomUserEditForm(instance=usuario)
    return render(request, "usuarios/editar.html", {"form": form, "usuario": usuario})


@login_required
@user_passes_test(is_admin)
def usuarios_eliminar(request, user_id):
    """Eliminar usuarios desde el listado (sólo por POST)."""
    usuario = get_object_or_404(User, id=user_id)

    if request.method != "POST":
        messages.error(request, "Acción no permitida.")
        return redirect("asistencias:usuarios_lista")

    # Evitar que el admin se elimine a sí mismo
    if request.user.id == usuario.id:
        messages.error(request, "No podés eliminar tu propio usuario.")
        return redirect("asistencias:usuarios_lista")

    # Evitar borrar el último superusuario del sistema
    if usuario.is_superuser and User.objects.filter(is_superuser=True).count() <= 1:
        messages.error(request, "No podés eliminar el último superusuario.")
        return redirect("asistencias:usuarios_lista")

    usuario.delete()
    messages.success(request, "Usuario eliminado correctamente.")
    return redirect("asistencias:usuarios_lista")


# =========================
# Académico: Carreras / Materias
# =========================
@login_required
@user_passes_test(is_admin)
def carreras_lista(request):
    carreras = Carrera.objects.all().order_by("nombre")
    return render(request, "admin/carreras_lista.html", {"carreras": carreras})


@login_required
@user_passes_test(is_admin)
def carrera_editar(request, carrera_id):
    from ..forms import CarreraForm
    carrera = get_object_or_404(Carrera, id=carrera_id)
    if request.method == "POST":
        form = CarreraForm(request.POST, instance=carrera)
        if form.is_valid():
            form.save()
            messages.success(request, "Carrera actualizada correctamente.")
            return redirect("asistencias:carreras_lista")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = CarreraForm(instance=carrera)
    return render(request, "admin/carrera_form.html", {"form": form, "carrera": carrera})


@login_required
@user_passes_test(is_admin)
def carrera_eliminar(request, carrera_id):
    carrera = get_object_or_404(Carrera, id=carrera_id)

    if request.method != "POST":
        messages.error(request, "Acción no permitida.")
        return redirect("asistencias:carreras_lista")

    carrera.delete()
    messages.success(request, "Carrera eliminada correctamente.")
    return redirect("asistencias:carreras_lista")


@login_required
@user_passes_test(is_admin)
def crear_carrera(request):
    from ..forms import CarreraForm
    if request.method == "POST":
        form = CarreraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Carrera creada correctamente.")
            return redirect("asistencias:carreras_lista")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = CarreraForm()
    return render(request, "admin/carrera_form.html", {"form": form, "carrera": None})


@login_required
@user_passes_test(is_admin)
def materias_lista(request):
    carrera_id = request.GET.get("carrera")
    materias = (
        Materia.objects
        .select_related("carrera")
        .all()
        .order_by("carrera__nombre", "nombre")
    )

    if carrera_id:
        materias = materias.filter(carrera_id=carrera_id)

    carreras = Carrera.objects.all().order_by("nombre")

    context = {
        "materias": materias,
        "carreras": carreras,
        "carrera_id": str(carrera_id or ""),
    }
    return render(request, "admin/materias_lista.html", context)


@login_required
@user_passes_test(is_admin)
def materia_editar(request, materia_id):
    from ..forms import MateriaForm
    materia = get_object_or_404(Materia, id=materia_id)
    if request.method == "POST":
        form = MateriaForm(request.POST, instance=materia)
        if form.is_valid():
            form.save()
            messages.success(request, "Materia actualizada correctamente.")
            return redirect("asistencias:materias_lista")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = MateriaForm(instance=materia)
    return render(request, "admin/materia_form.html", {"form": form, "materia": materia})


@login_required
@user_passes_test(is_admin)
def materia_eliminar(request, materia_id):
    materia = get_object_or_404(Materia, id=materia_id)

    if request.method != "POST":
        messages.error(request, "Acción no permitida.")
        return redirect("asistencias:materias_lista")

    materia.delete()
    messages.success(request, "Materia eliminada correctamente.")
    return redirect("asistencias:materias_lista")


@login_required
@user_passes_test(is_admin)
def crear_materia(request):
    from ..forms import MateriaForm
    if request.method == "POST":
        form = MateriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Materia creada correctamente.")
            return redirect("asistencias:materias_lista")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = MateriaForm()
    return render(request, "admin/materia_form.html", {"form": form, "materia": None})

# =========================
# Detalle de Carrera
# =========================
@login_required
@user_passes_test(is_admin)
def carrera_detalle(request, carrera_id):
    carrera = get_object_or_404(Carrera, id=carrera_id)

    # Materias de la carrera + cantidad de alumnos inscriptos por materia
    materias = (
        Materia.objects
        .filter(carrera=carrera)
        .annotate(
            total_alumnos=Count("alumnomateria__alumno", distinct=True)
        )
        .order_by("nombre")
    )

    # Alumnos inscriptos en cualquier materia de la carrera (sin repetir)
    alumnos = (
        Alumno.objects
        .filter(alumnomateria__materia__carrera=carrera)
        .distinct()
        .order_by("apellido", "nombre")
    )

    context = {
        "carrera": carrera,
        "materias": materias,
        "alumnos": alumnos,
    }
    return render(request, "admin/carrera_detalle.html", context)


# =========================
# Detalle de Materia
# =========================
@login_required
@user_passes_test(is_admin)
def materia_detalle(request, materia_id):
    materia = get_object_or_404(Materia, id=materia_id)

    # Cursadas (Docente + Período) de esa materia
    cursadas = (
        DocenteMateria.objects
        .filter(materia=materia)
        .select_related("docente", "periodo")
        .order_by("periodo__id")
    )

    # Para cada cursada, contamos alumnos inscriptos en ese período
    cursadas_datos = []
    for dm in cursadas:
        total_alumnos = AlumnoMateria.objects.filter(
            materia=materia,
            periodo=dm.periodo
        ).count()
        cursadas_datos.append({
            "cursada": dm,
            "total_alumnos": total_alumnos,
        })

    # Alumnos inscriptos en la materia (en cualquier período, sin repetir)
    alumnos = (
        Alumno.objects
        .filter(alumnomateria__materia=materia)
        .distinct()
        .order_by("apellido", "nombre")
    )

    context = {
        "materia": materia,
        "cursadas_datos": cursadas_datos,
        "alumnos": alumnos,
    }
    return render(request, "admin/materia_detalle.html", context)

# =========================
# Justificativos
# =========================
@login_required
@user_passes_test(is_admin)
def lista_justificativos(request):
    justificativos = (
        Asistencia.objects
        .filter(estado="Justificado")
        .select_related("alumno_materia__alumno", "alumno_materia__materia")
        .order_by("-fecha")
    )
    return render(request, "admin/justificativos.html", {"justificativos": justificativos})


@login_required
@user_passes_test(is_admin)
def aprobar_justificativo(request, asistencia_id):
    a = get_object_or_404(Asistencia, id=asistencia_id)
    # Si el ADMIN no es Docente, guardamos None para no romper la FK
    a.validado_por = get_object_or_404(Docente, user=request.user) if hasattr(request.user, "docente") else None
    a.validado_fecha = now()
    a.save()
    messages.success(request, "Justificativo aprobado correctamente.")
    return redirect("asistencias:lista_justificativos")


@login_required
@user_passes_test(is_admin)
def rechazar_justificativo(request, asistencia_id):
    a = get_object_or_404(Asistencia, id=asistencia_id)
    a.observaciones = (a.observaciones or "") + " | Rechazado por administración"
    a.estado = "Ausente"
    a.save()
    messages.error(request, "Justificativo rechazado.")
    return redirect("asistencias:lista_justificativos")


# =========================
# Reportes (filtro + paginación + export CSV/XLSX)
# =========================
@login_required
@user_passes_test(is_admin)
def reportes_curso(request):
    from pathlib import Path
    cursos = (
        DocenteMateria.objects
        .select_related("materia", "periodo", "docente")
        .all()
        .order_by("materia__nombre")
    )

    # Filtros
    curso_id = request.GET.get("curso", "")
    export = request.GET.get("export", "")  # "csv" o "xlsx"

    datos = []
    curso = None

    if curso_id:
        # curso seleccionado + inscriptos optimizados
        curso = (
            DocenteMateria.objects
            .select_related("materia", "periodo", "docente")
            .get(id=curso_id)
        )

        inscriptos = (
            AlumnoMateria.objects
            .filter(materia=curso.materia, periodo=curso.periodo)
            .select_related("alumno", "materia", "periodo")
            .prefetch_related("asistencia_set")
            .order_by("alumno__apellido", "alumno__nombre")
        )

        for am in inscriptos:
            qs = am.asistencia_set.all()
            total = len(qs)
            presentes = sum(1 for a in qs if a.estado == "Presente")
            justificados = sum(1 for a in qs if a.estado == "Justificado")
            ausentes = sum(1 for a in qs if a.estado in ["Ausente", "Tardanza"])
            porcentaje = round(((presentes + justificados) / total * 100), 2) if total else 0
            datos.append({
                "alumno": am.alumno,  # str(am.alumno) en export
                "total": total,
                "presentes": presentes,
                "justificados": justificados,
                "ausentes": ausentes,
                "porcentaje": porcentaje,  # 0..100
            })

    # ======== Exportar CSV simple ========
    if export == "csv" and curso and datos:
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="reporte_curso_{curso.id}.csv"'
        writer = csv.writer(response)
        writer.writerow(["Alumno", "Total", "Presentes", "Justificados", "Ausentes", "Porcentaje"])
        for d in datos:
            writer.writerow([
                str(d["alumno"]),
                d["total"],
                d["presentes"],
                d["justificados"],
                d["ausentes"],
                f'{d["porcentaje"]}%',
            ])
        return response

    # ======== Exportar XLSX con formato institucional ========
    if export == "xlsx" and curso and datos:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            messages.error(request, "Para exportar a XLSX instalá 'openpyxl' (pip install openpyxl).")
            return redirect(f"{request.path}?curso={curso.id}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"

        from pathlib import Path
        base_dir = Path(__file__).resolve().parents[2]
        logo_path = base_dir / "static" / "img" / "logo.png"
        start_row = 1
        if logo_path.exists():
            try:
                img = XLImage(str(logo_path))
                img.height = 60  # ajuste visual
                img.width = 60
                ws.add_image(img, "A1")
                start_row = 5  # dejamos espacio para el logo
            except Exception:
                start_row = 3

        # --- Título institucional ---
        titulo = (
            f"Reporte de Asistencia — {curso.materia.nombre} / "
            f"{curso.periodo.nombre} — Docente: {curso.docente}"
        )
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        cell_title = ws.cell(row=start_row, column=1, value=titulo)
        cell_title.font = Font(bold=True, size=14)
        cell_title.alignment = Alignment(horizontal="center")
        start_row += 2  # salteamos una fila

        # --- Encabezados
        headers = ["Alumno", "Total", "Presentes", "Justificados", "Ausentes", "% Asistencia"]
        header_row = start_row
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="111827")  # gris oscuro
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

        # --- Datos
        data_row_start = header_row + 1
        row = data_row_start
        for d in datos:
            ws.cell(row=row, column=1, value=str(d["alumno"]))
            ws.cell(row=row, column=2, value=d["total"])
            ws.cell(row=row, column=3, value=d["presentes"])
            ws.cell(row=row, column=4, value=d["justificados"])
            ws.cell(row=row, column=5, value=d["ausentes"])
            # porcentaje numérico (0..1) con formato 0.00%
            pcell = ws.cell(row=row, column=6, value=(d["porcentaje"] / 100.0))
            pcell.number_format = "0.00%"
            # bordes ligeros
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin", color="EEEEEE"),
                    right=Side(style="thin", color="EEEEEE"),
                    top=Side(style="thin", color="EEEEEE"),
                    bottom=Side(style="thin", color="EEEEEE"),
                )
            row += 1

        # --- Autoajuste de ancho de columnas (en base al contenido)
        from openpyxl.utils import get_column_letter
        for col in range(1, 7):
            col_letter = get_column_letter(col)
            max_len = 0
            for r in range(header_row, row):
                val = ws.cell(row=r, column=col).value
                if col == 6 and isinstance(val, (int, float)):
                    val_str = f"{val:.2%}"
                else:
                    val_str = str(val) if val is not None else ""
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 45)

        # --- Alineaciones
        from openpyxl.styles import Alignment
        for r in ws.iter_rows(min_row=data_row_start, min_col=2, max_col=6, max_row=row - 1):
            for c in r:
                c.alignment = Alignment(horizontal="center")

        # --- Congelar encabezado
        ws.freeze_panes = ws.cell(row=data_row_start, column=1)  # congela todo por encima

        # --- Resumen al final (opcional)
        summary_row = row + 1
        ws.cell(row=summary_row, column=1, value="Total alumnos:")
        ws.cell(row=summary_row, column=2, value=len(datos))
        ws.cell(row=summary_row, column=1).font = Font(bold=True)

        # --- Respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="reporte_curso_{curso.id}.xlsx"'
        wb.save(response)
        return response

    # ======== Paginación para la tabla HTML ========
    paginator = Paginator(datos, 15)  # 15 filas por página
    page_number = request.GET.get("page")
    datos_page = paginator.get_page(page_number)

    context = {
        "cursos": cursos,
        "curso": curso,
        "datos": datos_page,
        "curso_id": curso_id,
    }
    return render(request, "admin/reportes.html", context)


# =========================
# Cursadas / Asignaciones / Inscripciones
# =========================
@login_required
@user_passes_test(is_admin)
def admin_cursadas(request):
    cursadas = (
        DocenteMateria.objects
        .select_related("materia", "periodo", "docente", "docente__user")
        .order_by("materia__nombre", "periodo__id")
    )
    return render(request, "admin/cursadas.html", {"cursadas": cursadas})

# =========================
# Detalle de cursada
# =========================
@login_required
@user_passes_test(is_admin)
def cursada_detalle(request, cursada_id):
    """Detalle de una cursada (DocenteMateria) con estadísticas de asistencia."""
    from django.shortcuts import get_object_or_404
    from django.db.models import Prefetch

    dm = get_object_or_404(
        DocenteMateria.objects.select_related("materia", "periodo", "docente"),
        id=cursada_id,
    )

    # Alumnos inscriptos en esa materia/período, con sus asistencias prefetchadas
    inscriptos = (
        AlumnoMateria.objects
        .filter(materia=dm.materia, periodo=dm.periodo)
        .select_related("alumno", "alumno__user")
        .prefetch_related("asistencia_set")
        .order_by("alumno__apellido", "alumno__nombre")
    )

    datos = []
    total_registros = 0
    total_ok = 0  # presentes + justificados

    for am in inscriptos:
        qs = list(am.asistencia_set.all())
        total = len(qs)
        presentes = sum(1 for a in qs if a.estado == "Presente")
        justificados = sum(1 for a in qs if a.estado == "Justificado")
        ausentes = sum(1 for a in qs if a.estado in ["Ausente", "Tardanza"])
        porcentaje = round(((presentes + justificados) / total * 100), 2) if total else 0

        datos.append({
            "alumno": am.alumno,
            "dni": getattr(am.alumno, "dni", ""),
            "total": total,
            "presentes": presentes,
            "justificados": justificados,
            "ausentes": ausentes,
            "porcentaje": porcentaje,
        })

        total_registros += total
        total_ok += (presentes + justificados)

    total_alumnos = len(datos)
    porcentaje_global = round((total_ok / total_registros * 100), 2) if total_registros else 0

    # ===== Exportar a Excel (XLSX) =====
    export = request.GET.get("export")
    if export == "xlsx" and datos:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messages.error(request, "Para exportar a Excel instalá 'openpyxl' (pip install openpyxl).")
            return redirect(request.path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistencia cursada"

        # Título
        titulo = (
            f"Asistencia — {dm.materia.nombre} / {dm.periodo.nombre} "
            f"(Docente: {dm.docente})"
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        cell_title = ws.cell(row=1, column=1, value=titulo)
        cell_title.font = Font(bold=True, size=14)
        cell_title.alignment = Alignment(horizontal="center")

        # Encabezados
        headers = ["Alumno", "DNI", "Total", "Presentes", "Justificados", "Ausentes", "% Asistencia"]
        header_row = 3
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="111827")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

        # Datos
        row = header_row + 1
        for d in datos:
            ws.cell(row=row, column=1, value=str(d["alumno"]))
            ws.cell(row=row, column=2, value=d["dni"])
            ws.cell(row=row, column=3, value=d["total"])
            ws.cell(row=row, column=4, value=d["presentes"])
            ws.cell(row=row, column=5, value=d["justificados"])
            ws.cell(row=row, column=6, value=d["ausentes"])
            pcell = ws.cell(row=row, column=7, value=(d["porcentaje"] / 100.0))
            pcell.number_format = "0.00%"

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin", color="EEEEEE"),
                    right=Side(style="thin", color="EEEEEE"),
                    top=Side(style="thin", color="EEEEEE"),
                    bottom=Side(style="thin", color="EEEEEE"),
                )
            row += 1

        # Autoajuste de columnas
        for col in range(1, 8):
            col_letter = get_column_letter(col)
            max_len = 0
            for r in range(1, row):
                val = ws.cell(row=r, column=col).value
                if col == 7 and isinstance(val, (int, float)):
                    val_str = f"{val:.2%}"
                else:
                    val_str = str(val) if val is not None else ""
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 45)

        # Resumen al final
        summary_row = row + 1
        ws.cell(row=summary_row, column=1, value="Total alumnos:")
        ws.cell(row=summary_row, column=2, value=total_alumnos)
        ws.cell(row=summary_row + 1, column=1, value="% asistencia global:")
        ws.cell(row=summary_row + 1, column=2, value=(porcentaje_global / 100.0)).number_format = "0.00%"

        # Respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"asistencia_cursada_{dm.id}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    context = {
        "cursada": dm,
        "inscriptos": inscriptos,
        "datos": datos,
        "total_alumnos": total_alumnos,
        "porcentaje_global": porcentaje_global,
    }
    return render(request, "admin/cursada_detalle.html", context)

@login_required
@user_passes_test(is_admin)
def asignar_docente(request):
    from ..forms import DocenteMateriaForm
    if request.method == "POST":
        form = DocenteMateriaForm(request.POST)
        if form.is_valid():
            dm = form.save()
            messages.success(request, f"Asignado {dm.docente} a {dm.materia} ({dm.periodo}).")
            return redirect("asistencias:admin_cursadas")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = DocenteMateriaForm()
    return render(request, "admin/asignar_docente.html", {"form": form})


@login_required
@user_passes_test(is_admin)
def inscribir_alumnos(request):
    from ..forms import AlumnoMateriaForm
    if request.method == "POST":
        form = AlumnoMateriaForm(request.POST)
        if form.is_valid():
            materia = form.cleaned_data["materia"]
            periodo = form.cleaned_data["periodo"]
            alumnos = form.cleaned_data["alumnos"]
            creados = 0
            for a in alumnos:
                _, created = AlumnoMateria.objects.get_or_create(
                    alumno=a,
                    materia=materia,
                    periodo=periodo,
                )
                if created:
                    creados += 1
            messages.success(request, f"Inscripciones registradas: {creados}.")
            return redirect("asistencias:admin_cursadas")
        messages.error(request, "Revisá los datos del formulario.")
    else:
        form = AlumnoMateriaForm()
    return render(request, "admin/inscribir_alumnos.html", {"form": form})


# =========================
# Métricas (Admin / Docente / Alumno)
# =========================
@login_required
@user_passes_test(is_admin)
def admin_metricas(request):
    # Filtros opcionales
    materia_id = request.GET.get("materia")
    periodo_id = request.GET.get("periodo")

    asistencias = (
        Asistencia.objects
        .select_related("alumno_materia__materia", "alumno_materia__periodo")
    )

    if materia_id:
        asistencias = asistencias.filter(alumno_materia__materia_id=materia_id)
    if periodo_id:
        asistencias = asistencias.filter(alumno_materia__periodo_id=periodo_id)

    # KPIs
    agg = asistencias.aggregate(
        total=Count("id"),
        presentes=Count(Case(When(estado="Presente", then=1), output_field=IntegerField())),
        justificados=Count(Case(When(estado="Justificado", then=1), output_field=IntegerField())),
        ausentes=Count(Case(When(estado="Ausente", then=1), output_field=IntegerField())),
        tardanzas=Count(Case(When(estado="Tardanza", then=1), output_field=IntegerField())),
    )
    total = agg["total"] or 0
    pct = lambda x: round((x / total * 100), 2) if total else 0
    kpis = {
        "total": total,
        "presentes": agg["presentes"] or 0,
        "justificados": agg["justificados"] or 0,
        "ausentes": agg["ausentes"] or 0,
        "tardanzas": agg["tardanzas"] or 0,
        "porcentaje_asistencia": pct((agg["presentes"] or 0) + (agg["justificados"] or 0)),
    }

    # Por materia (para gráfica)
    por_materia = (
        asistencias
        .values("alumno_materia__materia__id", "alumno_materia__materia__nombre")
        .annotate(
            total=Count("id"),
            ok=Count(Case(
                When(estado__in=["Presente", "Justificado"], then=1),
                output_field=IntegerField()
            )),
        )
        .order_by("alumno_materia__materia__nombre")
    )

    chart_labels = [r["alumno_materia__materia__nombre"] for r in por_materia]
    chart_values = [
        round((r["ok"] / r["total"] * 100), 2) if r["total"] else 0
        for r in por_materia
    ]

    materias = Materia.objects.all().order_by("nombre")
    periodos = Periodo.objects.all().order_by("-id")

    context = {
        "kpis": kpis,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "materias": materias,
        "periodos": periodos,
        "materia_id": str(materia_id or ""),
        "periodo_id": str(periodo_id or ""),
    }
    return render(request, "admin/metricas.html", context)

@login_required
def alumno_metricas(request):
    # Asegurar rol alumno
    if not hasattr(request.user, "alumno"):
        messages.error(request, "Acceso restringido a alumnos.")
        return redirect("asistencias:home")

    alumno = request.user.alumno
    asistencias = (
        Asistencia.objects
        .filter(alumno_materia__alumno=alumno)
        .select_related("alumno_materia__materia", "alumno_materia__periodo")
    )

    agg = asistencias.aggregate(
        total=Count("id"),
        presentes=Count(Case(When(estado="Presente", then=1), output_field=IntegerField())),
        justificados=Count(Case(When(estado="Justificado", then=1), output_field=IntegerField())),
        ausentes=Count(Case(When(estado="Ausente", then=1), output_field=IntegerField())),
        tardanzas=Count(Case(When(estado="Tardanza", then=1), output_field=IntegerField())),
    )
    total = agg["total"] or 0
    pct = lambda x: round((x / total * 100), 2) if total else 0
    kpis = {
        "total": total,
        "presentes": agg["presentes"] or 0,
        "justificados": agg["justificados"] or 0,
        "ausentes": agg["ausentes"] or 0,
        "tardanzas": agg["tardanzas"] or 0,
        "porcentaje_asistencia": pct((agg["presentes"] or 0) + (agg["justificados"] or 0)),
    }

    por_materia = (
        asistencias
        .values("alumno_materia__materia__nombre")
        .annotate(
            total=Count("id"),
            ok=Count(Case(
                When(estado__in=["Presente", "Justificado"], then=1),
                output_field=IntegerField()
            ))
        )
        .order_by("alumno_materia__materia__nombre")
    )

    chart_labels = [r["alumno_materia__materia__nombre"] for r in por_materia]
    chart_values = [
        round((r["ok"] / r["total"] * 100), 2) if r["total"] else 0
        for r in por_materia
    ]

    context = {
        "kpis": kpis,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
    }
    return render(request, "alumno/metricas.html", context)
